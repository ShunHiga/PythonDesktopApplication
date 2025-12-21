import FreeSimpleGUI as sg
import openpyxl as xl

#Excelファイルから人口統計情報を読み込む関数
def read_population():
	print("read_population")
	#Excelファイルを読み込む
	EXCEL_FILE = "./population_jp.xlsx"
	workbook = xl.load_workbook(EXCEL_FILE)
	#ブックの中のシート「A」を取得
	sheet = workbook["A"]
	#都道府県ごとの総人口を得る
	#メモ：都道府県名はI14からI60のセル
	#総人口はL14からL60のセル
	result = []

	#セル名から列番号を得る
	name_col_index = xl.utils.column_index_from_string("I")
	pop_col_index = xl.utils.column_index_from_string("L")

	#連続でセルの値を取得する
	for row in range(14, 60 + 1):
		name = sheet.cell(row, name_col_index).value
		pop = sheet.cell(row, pop_col_index).value
		man = sheet.cell(row, pop_col_index + 2).value
		woman = sheet.cell(row, pop_col_index + 4).value
		result.append([name, pop, man,woman])
	
	workbook.close()
	return result

#テーブルに結果を表示する関数
def show_table(data):
	layout = [
		[sg.Table(values=data,
			headings=["都道府県","総人口(万人)","男性","女性"],
			auto_size_columns=True,
			expand_x=True,expand_y=True,
			justification="right",
			font=("Arial",14))],
		[sg.Button("閉じる")]
	]
	#ウィンドウを作成
	window = sg.Window("人口統計",layout, size=(600,400))
	while True:
		event, _ = window.read()
		if event in [sg.WIN_CLOSED, "閉じる"]:
			break
	window.close()

if __name__== "__main__":
	#メイン処理
	data = read_population()#人口統計情報を読み込む
	show_table(data)
	print("テスト")