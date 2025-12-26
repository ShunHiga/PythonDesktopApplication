import os
from xhtml2pdf import pisa
import openpyxl as xl

#顧客名簿を読み込む
workbook = xl.load_workbook("./meibo.xlsx")
sheet = workbook.active

#招待状のひな形となるHTMLファイルを読む
with open("invitation.html", "r", encoding="utf-8") as f:
	template = f.read()

#セルの内容を連続で読む
for i in range(2, 9999):
	#i行目の情報を得る
	id = sheet.cell(row=i, column=1).value
	name = sheet.cell(row=i, column=2).value
	email = sheet.cell(row=i, column=3).value
	no = sheet.cell(row=i, column=4).value
	if id is None:
		break

	#招待状のHTMLを生成
	html = template.replace('__name__', name)
	html = html.replace('__no__', str(no))
	html = html.replace('__email__', email)

	#PDFを出力
	if os.path.exists("./invitation") == False:
		os.mkdir("./invitation")
	
	with open(f"./invitation/{id}.pdf","wb") as pdf_file:
		pisa.CreatePDF(html, dest=pdf_file)
		print(f"{id}:{name}さんの招待状を作成しました")